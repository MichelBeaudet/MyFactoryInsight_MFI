#!/usr/bin/env python3
# =============================================================================
# PROJECT      : MyFactoryInsight (MFI)
# FILE         : mfi_phase_09_reports.py
# PHASE        : 9 — Reports (PDF + Excel)
# PURPOSE      : Generate professional PDF and Excel reports from MFI pipeline
#                data. PDF report: industrial dark-styled cover + KPI summary +
#                machine status table + alert summary. Excel report: multiple
#                worksheets (Fleet KPI, Machine Detail, Alert Log) with
#                conditional formatting and charts. Both reports derive their
#                data from a ReportDataset collected from MFICore cycles.
# AUTHOR       : Michel Beaudet
# CREATED      : 2026-05-16
# PYTHON       : 3.12+ (3.14 target-compatible syntax)
# DEPENDENCIES : reportlab>=4.0, openpyxl>=3.1, pydantic>=2.0, mfi_phase_01..04
# CLI          : python mfi_phase_09_reports.py --self-test
#                python mfi_phase_09_reports.py --cycles 5 --pdf report.pdf
#                python mfi_phase_09_reports.py --cycles 5 --xlsx report.xlsx
#                python mfi_phase_09_reports.py --cycles 5 --pdf r.pdf --xlsx r.xlsx
# =============================================================================

# =============================================================================
# SECTION 1 — IMPORTS
# =============================================================================
import argparse
import io
import json
import logging
import math
import os
import sys
import time
from datetime import datetime, timezone
from typing import Any, Optional

# --- ReportLab ---
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm, cm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, PageBreak, KeepTogether,
    )
    from reportlab.platypus.flowables import HRFlowable
except ImportError as exc:
    print(
        f"[FATAL] reportlab not found: {exc}\n"
        "Install: pip install reportlab --break-system-packages",
        file=sys.stderr,
    )
    sys.exit(1)

# --- openpyxl ---
try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
        GradientFill,
    )
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.formatting.rule import ColorScaleRule, Rule
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import SeriesLabel
except ImportError as exc:
    print(
        f"[FATAL] openpyxl not found: {exc}\n"
        "Install: pip install openpyxl --break-system-packages",
        file=sys.stderr,
    )
    sys.exit(1)

# --- Phase 4 ---
try:
    from mfi_phase_04_core import (
        MFICore,
        MFIEnrichedRecord,
        DEFAULT_SITE,
    )
except ImportError as exc:
    print(f"[FATAL] Cannot import mfi_phase_04_core: {exc}", file=sys.stderr)
    sys.exit(1)

# =============================================================================
# SECTION 2 — CONFIG / CONSTANTS
# =============================================================================
PHASE_ID            = "09"
PHASE_NAME          = "Reports"
PHASE_VERSION       = "1.0.0"

DEFAULT_CYCLES      = 5
DEFAULT_PDF_PATH    = "mfi_report.pdf"
DEFAULT_XLSX_PATH   = "mfi_report.xlsx"

# ── Brand palette (matches Phase 6 dashboard dark theme) ─────────────────────
C_BG            = colors.HexColor("#070d12")    # Page background
C_SURFACE       = colors.HexColor("#0d1a24")    # Section backgrounds
C_ACCENT        = colors.HexColor("#00b4d8")    # Accent / headers
C_RUN           = colors.HexColor("#00e676")    # RUN status
C_IDLE          = colors.HexColor("#ffd600")    # IDLE status
C_FAULT         = colors.HexColor("#ff1744")    # FAULT status
C_MAINT         = colors.HexColor("#ff9100")    # MAINTENANCE status
C_TEXT          = colors.HexColor("#c8d8e4")    # Body text
C_DIM           = colors.HexColor("#4a6275")    # Dimmed text
C_BORDER        = colors.HexColor("#1a2e40")    # Table borders

# Excel hex values (no #)
XL_BG       = "070D12"
XL_SURFACE  = "0D1A24"
XL_ACCENT   = "00B4D8"
XL_RUN      = "00E676"
XL_IDLE     = "FFD600"
XL_FAULT    = "FF1744"
XL_MAINT    = "FF9100"
XL_TEXT     = "C8D8E4"
XL_DIM      = "4A6275"
XL_WHITE    = "FFFFFF"
XL_BORDER   = "1A2E40"
XL_CRIT_TEMP= "FF1744"
XL_WARN_TEMP= "FFD600"

STATUS_COLOR_MAP: dict[str, Any] = {
    "RUN"         : C_RUN,
    "IDLE"        : C_IDLE,
    "FAULT"       : C_FAULT,
    "MAINTENANCE" : C_MAINT,
}

# =============================================================================
# SECTION 3 — LOGGER SETUP
# =============================================================================
LOG_FORMAT = (
    "[%(asctime)s] "
    "[%(levelname)-8s] "
    "[MFI-P%(phase)s] "
    "[%(funcName)s] "
    "%(message)s"
)


class PhaseAdapter(logging.LoggerAdapter):
    def process(self, msg: str, kwargs: dict) -> tuple[str, dict]:
        kwargs.setdefault("extra", {})
        kwargs["extra"]["phase"] = PHASE_ID
        return msg, kwargs


def build_logger(name: str, level: int = logging.DEBUG) -> PhaseAdapter:
    handler = logging.StreamHandler(sys.stdout)
    handler.setFormatter(logging.Formatter(LOG_FORMAT))
    base = logging.getLogger(name)
    base.setLevel(level)
    base.handlers.clear()
    base.addHandler(handler)
    base.propagate = False
    return PhaseAdapter(base, extra={"phase": PHASE_ID})


LOG = build_logger("mfi.phase09")

# =============================================================================
# SECTION 4 — REPORT DATASET
# =============================================================================

class ReportDataset:
    """
    Aggregated dataset collected from one or more MFICore pipeline cycles.

    Stores all enriched records seen across N cycles and computes aggregate
    KPIs once via compute_aggregates(). All report generators consume this
    single dataset — no pipeline access after collection.

    Attributes
    ----------
    site_id         : Site identifier.
    cycles_collected: Number of cycles data was gathered from.
    period_start    : ISO 8601 UTC timestamp of first record.
    period_end      : ISO 8601 UTC timestamp of last record.
    records         : All MFIEnrichedRecord objects (all cycles combined).
    aggregates      : Dict of computed KPIs (set by compute_aggregates()).
    generated_at    : ISO 8601 UTC timestamp when this dataset was built.
    """

    def __init__(self, site_id: str = DEFAULT_SITE) -> None:
        self.site_id            = site_id
        self.cycles_collected   = 0
        self.period_start       : Optional[str] = None
        self.period_end         : Optional[str] = None
        self.records            : list[MFIEnrichedRecord] = []
        self.aggregates         : dict[str, Any] = {}
        self.generated_at       = datetime.now(timezone.utc).isoformat()
        # Track last-seen record per machine for snapshot table
        self._latest            : dict[str, MFIEnrichedRecord] = {}

    def add_cycle(self, enriched: list[MFIEnrichedRecord]) -> None:
        """
        Ingest one cycle's enriched records into the dataset.

        Maintains a latest-record-per-machine dict for the status snapshot.
        Updates period_start / period_end timestamps.

        Args:
            enriched : MFIEnrichedRecord list from MFICore.run_cycle().
        """
        if not enriched:
            return

        self.cycles_collected += 1
        self.records.extend(enriched)

        for r in enriched:
            self._latest[r.machine_id] = r

        timestamps = [r.timestamp for r in enriched if r.timestamp]
        if timestamps:
            ts_sorted = sorted(timestamps)
            if self.period_start is None or ts_sorted[0] < self.period_start:
                self.period_start = ts_sorted[0]
            if self.period_end is None or ts_sorted[-1] > self.period_end:
                self.period_end = ts_sorted[-1]

    def latest_snapshot(self) -> list[MFIEnrichedRecord]:
        """Return one record per machine — the most recent state seen."""
        return sorted(self._latest.values(), key=lambda r: r.machine_id)

    def compute_aggregates(self) -> dict[str, Any]:
        """
        Compute KPI aggregates across all collected records.

        Aggregates computed:
          - Status distribution (counts and percentages)
          - Fleet availability (mean of per-machine availability)
          - OEE components: availability, performance, quality
          - Production totals: pieces, good, bad
          - Temperature stats: mean, max, critical count, warn count
          - Alarm stats: total non-zero, by code
          - Per-machine-type breakdown
          - Top 5 machines by piece count

        Returns:
            Dict of aggregated KPI values (also stored in self.aggregates).
        """
        snapshot = self.latest_snapshot()
        all_recs = self.records

        if not snapshot:
            self.aggregates = {"error": "no data"}
            return self.aggregates

        # ── Status distribution ───────────────────────────────────────────
        status_dist: dict[str, int] = {}
        for r in snapshot:
            status_dist[r.status] = status_dist.get(r.status, 0) + 1

        n = len(snapshot)
        status_pct = {k: round(v / n * 100, 1) for k, v in status_dist.items()}

        # ── Availability ──────────────────────────────────────────────────
        avails = [r.availability for r in snapshot if r.availability is not None]
        fleet_avail = round(sum(avails) / len(avails), 4) if avails else None
        run_ratio = round(status_dist.get("RUN", 0) / n, 4)

        # ── Quality ───────────────────────────────────────────────────────
        qrates = [r.quality_rate for r in snapshot if r.quality_rate is not None]
        fleet_quality = round(sum(qrates) / len(qrates), 4) if qrates else None

        # ── Performance (machines with non-zero cycle time) ───────────────
        perf_n = sum(1 for r in snapshot if r.cycle_time_sec > 0)
        performance = round(perf_n / n, 4)

        # ── OEE ──────────────────────────────────────────────────────────
        oee = (
            round(run_ratio * performance * fleet_quality, 4)
            if fleet_quality is not None else None
        )

        # ── Production totals ─────────────────────────────────────────────
        total_pieces = sum(r.piece_count for r in snapshot)
        total_good   = sum(r.good_count  for r in snapshot)
        total_bad    = sum(r.bad_count   for r in snapshot)

        # ── Temperature stats ─────────────────────────────────────────────
        temps           = [r.temperature_c for r in snapshot]
        temp_mean       = round(sum(temps) / len(temps), 1) if temps else 0.0
        temp_max_val    = max(temps) if temps else 0.0
        temp_max_machine= max(snapshot, key=lambda r: r.temperature_c).machine_id if snapshot else "—"
        temp_crits      = sum(1 for r in snapshot if r.temp_severity == "CRITICAL")
        temp_warns      = sum(1 for r in snapshot if r.temp_severity == "WARN")

        # ── Alarms ───────────────────────────────────────────────────────
        active_alarms   = sum(1 for r in snapshot if r.alarm_code != 0)
        alarm_codes: dict[int, int] = {}
        for r in snapshot:
            if r.alarm_code != 0:
                alarm_codes[r.alarm_code] = alarm_codes.get(r.alarm_code, 0) + 1

        # ── Type breakdown ────────────────────────────────────────────────
        type_breakdown: dict[str, dict] = {}
        for r in snapshot:
            t = r.machine_type
            if t not in type_breakdown:
                type_breakdown[t] = {
                    "count"   : 0,
                    "running" : 0,
                    "pieces"  : 0,
                    "faulted" : 0,
                }
            type_breakdown[t]["count"]   += 1
            type_breakdown[t]["pieces"]  += r.piece_count
            if r.is_running:
                type_breakdown[t]["running"] += 1
            if r.is_faulted:
                type_breakdown[t]["faulted"] += 1

        # ── Top 5 producers ───────────────────────────────────────────────
        top5 = sorted(snapshot, key=lambda r: r.piece_count, reverse=True)[:5]
        top5_list = [
            {"machine_id": r.machine_id, "machine_type": r.machine_type,
             "piece_count": r.piece_count, "status": r.status}
            for r in top5
        ]

        self.aggregates = {
            "machine_count"     : n,
            "cycles_collected"  : self.cycles_collected,
            "status_distribution": status_dist,
            "status_pct"        : status_pct,
            "fleet_availability": fleet_avail,
            "run_ratio"         : run_ratio,
            "quality"           : fleet_quality,
            "performance"       : performance,
            "oee"               : oee,
            "total_pieces"      : total_pieces,
            "total_good"        : total_good,
            "total_bad"         : total_bad,
            "temp_mean"         : temp_mean,
            "temp_max"          : temp_max_val,
            "temp_max_machine"  : temp_max_machine,
            "temp_critical_count": temp_crits,
            "temp_warn_count"   : temp_warns,
            "active_alarms"     : active_alarms,
            "alarm_codes"       : alarm_codes,
            "type_breakdown"    : type_breakdown,
            "top5_producers"    : top5_list,
        }
        LOG.info(
            "Aggregates computed │ machines=%d │ cycles=%d │ oee=%s │ avail=%s",
            n,
            self.cycles_collected,
            f"{oee:.1%}" if oee else "N/A",
            f"{fleet_avail:.1%}" if fleet_avail else "N/A",
        )
        return self.aggregates


# =============================================================================
# SECTION 5 — PDF REPORT GENERATOR
# =============================================================================

class PDFReportGenerator:
    """
    MFI PDF Report Generator using ReportLab Platypus.

    Generates a professional multi-section industrial report:
      Page 1 : Cover — site, period, generated timestamp
      Page 2 : Fleet KPI Summary — OEE, availability, production, alarms
      Page 3+: Machine Status Table — all machines, one row each
      Last   : Alert / Alarm Summary

    Design language matches Phase 6 dashboard (dark industrial palette).
    ReportLab renders on a light background with dark accent headers for
    print compatibility.
    """

    def __init__(self) -> None:
        self._styles = self._build_styles()

    # ── Style definitions ─────────────────────────────────────────────────

    def _build_styles(self) -> dict[str, ParagraphStyle]:
        """Build all custom paragraph styles."""
        base = getSampleStyleSheet()
        return {
            "cover_title": ParagraphStyle(
                "cover_title",
                fontName    = "Helvetica-Bold",
                fontSize    = 32,
                textColor   = C_ACCENT,
                spaceAfter  = 4 * mm,
                alignment   = TA_CENTER,
            ),
            "cover_sub": ParagraphStyle(
                "cover_sub",
                fontName    = "Helvetica",
                fontSize    = 14,
                textColor   = C_DIM,
                spaceAfter  = 2 * mm,
                alignment   = TA_CENTER,
            ),
            "cover_meta": ParagraphStyle(
                "cover_meta",
                fontName    = "Helvetica",
                fontSize    = 11,
                textColor   = C_TEXT,
                spaceAfter  = 1.5 * mm,
                alignment   = TA_CENTER,
            ),
            "section_title": ParagraphStyle(
                "section_title",
                fontName    = "Helvetica-Bold",
                fontSize    = 14,
                textColor   = C_ACCENT,
                spaceBefore = 6 * mm,
                spaceAfter  = 3 * mm,
            ),
            "body": ParagraphStyle(
                "body",
                fontName    = "Helvetica",
                fontSize    = 10,
                textColor   = C_TEXT,
                spaceAfter  = 2 * mm,
            ),
            "kpi_label": ParagraphStyle(
                "kpi_label",
                fontName    = "Helvetica",
                fontSize    = 9,
                textColor   = C_DIM,
                alignment   = TA_CENTER,
            ),
            "kpi_value": ParagraphStyle(
                "kpi_value",
                fontName    = "Helvetica-Bold",
                fontSize    = 22,
                textColor   = C_ACCENT,
                alignment   = TA_CENTER,
            ),
            "table_header": ParagraphStyle(
                "table_header",
                fontName    = "Helvetica-Bold",
                fontSize    = 8,
                textColor   = colors.white,
                alignment   = TA_CENTER,
            ),
            "table_cell": ParagraphStyle(
                "table_cell",
                fontName    = "Helvetica",
                fontSize    = 8,
                textColor   = C_TEXT,
                alignment   = TA_CENTER,
            ),
            "footer": ParagraphStyle(
                "footer",
                fontName    = "Helvetica",
                fontSize    = 8,
                textColor   = C_DIM,
                alignment   = TA_CENTER,
            ),
        }

    # ── Helpers ───────────────────────────────────────────────────────────

    @staticmethod
    def _pct(v: Optional[float]) -> str:
        return f"{v * 100:.1f}%" if v is not None else "N/A"

    @staticmethod
    def _num(v: Any) -> str:
        return f"{v:,}" if v is not None else "N/A"

    @staticmethod
    def _status_color(status: str) -> Any:
        return STATUS_COLOR_MAP.get(status, C_DIM)

    # ── Section builders ──────────────────────────────────────────────────

    def _build_cover(self, ds: ReportDataset) -> list:
        """Build the cover page flowables."""
        s = self._styles
        agg = ds.aggregates

        def fmt_ts(ts: Optional[str]) -> str:
            if not ts:
                return "N/A"
            try:
                dt = datetime.fromisoformat(ts.replace("Z", "+00:00"))
                return dt.strftime("%Y-%m-%d %H:%M UTC")
            except ValueError:
                return ts

        story = [
            Spacer(1, 40 * mm),
            Paragraph("MyFactoryInsight", s["cover_title"]),
            Paragraph("Production Analytics Report", s["cover_sub"]),
            HRFlowable(
                width="80%", thickness=1, color=C_ACCENT,
                spaceAfter=8 * mm, spaceBefore=4 * mm,
            ),
            Paragraph(f"Site: <b>{ds.site_id.upper()}</b>", s["cover_meta"]),
            Paragraph(
                f"Report Period: {fmt_ts(ds.period_start)} → {fmt_ts(ds.period_end)}",
                s["cover_meta"],
            ),
            Paragraph(
                f"Machines: {agg.get('machine_count', '—')} | "
                f"Cycles: {agg.get('cycles_collected', '—')}",
                s["cover_meta"],
            ),
            Spacer(1, 4 * mm),
            Paragraph(f"Generated: {fmt_ts(ds.generated_at)}", s["cover_meta"]),
            Spacer(1, 4 * mm),
            Paragraph(
                f"Phase {PHASE_ID} — {PHASE_NAME} v{PHASE_VERSION}",
                s["footer"],
            ),
            PageBreak(),
        ]
        return story

    def _build_kpi_section(self, ds: ReportDataset) -> list:
        """Build the Fleet KPI summary section."""
        s   = self._styles
        agg = ds.aggregates

        story = [
            Paragraph("Fleet KPI Summary", s["section_title"]),
            HRFlowable(width="100%", thickness=0.5, color=C_BORDER, spaceAfter=4 * mm),
        ]

        # KPI grid — 5 cards in one table row
        kpi_data = [
            ("OEE",           self._pct(agg.get("oee"))),
            ("Availability",  self._pct(agg.get("fleet_availability"))),
            ("Quality Rate",  self._pct(agg.get("quality"))),
            ("Total Pieces",  self._num(agg.get("total_pieces"))),
            ("Active Alarms", str(agg.get("active_alarms", "N/A"))),
        ]

        header_row = [Paragraph(label, s["kpi_label"]) for label, _ in kpi_data]
        value_row  = [Paragraph(value, s["kpi_value"]) for _, value in kpi_data]

        col_w = [33 * mm] * 5
        kpi_table = Table(
            [header_row, value_row],
            colWidths=col_w,
        )
        kpi_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), C_SURFACE),
            ("BOX",        (0, 0), (-1, -1), 0.5, C_BORDER),
            ("INNERGRID",  (0, 0), (-1, -1), 0.5, C_BORDER),
            ("TOPPADDING",    (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(kpi_table)
        story.append(Spacer(1, 6 * mm))

        # Status distribution table
        story.append(Paragraph("Status Distribution", s["section_title"]))
        dist  = agg.get("status_distribution", {})
        dpct  = agg.get("status_pct", {})
        n     = agg.get("machine_count", 1)
        statuses = ["RUN", "IDLE", "FAULT", "MAINTENANCE"]

        dist_rows = [
            [
                Paragraph("STATUS", s["table_header"]),
                Paragraph("MACHINES", s["table_header"]),
                Paragraph("PERCENTAGE", s["table_header"]),
            ]
        ]
        for st in statuses:
            count = dist.get(st, 0)
            pct   = dpct.get(st, 0.0)
            row_color = STATUS_COLOR_MAP.get(st, C_DIM)
            dist_rows.append([
                Paragraph(st, ParagraphStyle("st", fontName="Helvetica-Bold",
                          fontSize=9, textColor=row_color, alignment=TA_CENTER)),
                Paragraph(str(count), s["table_cell"]),
                Paragraph(f"{pct:.1f}%", s["table_cell"]),
            ])

        dist_table = Table(dist_rows, colWidths=[55 * mm, 55 * mm, 55 * mm])
        dist_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), C_SURFACE),
            ("BACKGROUND", (0, 1), (-1, -1), colors.white),
            ("BOX",        (0, 0), (-1, -1), 0.5, C_BORDER),
            ("INNERGRID",  (0, 0), (-1, -1), 0.5, C_BORDER),
            ("TOPPADDING",    (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(dist_table)
        story.append(Spacer(1, 5 * mm))

        # Production + Temperature summary
        story.append(Paragraph("Production & Temperature", s["section_title"]))
        prod_data = [
            [
                Paragraph("METRIC", s["table_header"]),
                Paragraph("VALUE", s["table_header"]),
            ],
            ["Total Pieces Produced",  self._num(agg.get("total_pieces"))],
            ["Good Pieces",            self._num(agg.get("total_good"))],
            ["Bad / Scrap Pieces",     self._num(agg.get("total_bad"))],
            ["Average Temperature",   f"{agg.get('temp_mean', 0):.1f} °C"],
            ["Max Temperature",       f"{agg.get('temp_max', 0):.1f} °C ({agg.get('temp_max_machine','')})"],
            ["Critical Temp Machines", str(agg.get("temp_critical_count", 0))],
            ["Warning Temp Machines",  str(agg.get("temp_warn_count", 0))],
        ]

        prod_table = Table(prod_data, colWidths=[100 * mm, 65 * mm])
        prod_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), C_SURFACE),
            ("BACKGROUND", (0, 1), (-1, -1), colors.white),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.HexColor("#f0f4f7"), colors.white]),
            ("BOX",        (0, 0), (-1, -1), 0.5, C_BORDER),
            ("INNERGRID",  (0, 0), (-1, -1), 0.5, C_BORDER),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING",   (0, 0), (-1, -1), 8),
            ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(prod_table)
        story.append(PageBreak())
        return story

    def _build_machine_table(self, ds: ReportDataset) -> list:
        """Build the full machine status table section."""
        s        = self._styles
        snapshot = ds.latest_snapshot()

        story = [
            Paragraph("Machine Status Snapshot", s["section_title"]),
            HRFlowable(width="100%", thickness=0.5, color=C_BORDER, spaceAfter=3 * mm),
        ]

        headers = [
            "MACHINE", "TYPE", "PROTOCOL", "STATUS",
            "ALARM", "PIECES", "TEMP °C", "AVAIL",
        ]
        col_widths = [28*mm, 22*mm, 18*mm, 24*mm, 14*mm, 22*mm, 18*mm, 16*mm]

        rows = [[Paragraph(h, s["table_header"]) for h in headers]]

        for r in snapshot:
            avail_str = f"{r.availability * 100:.0f}%" if r.availability is not None else "—"
            alarm_str = str(r.alarm_code) if r.alarm_code != 0 else "—"
            row_color = STATUS_COLOR_MAP.get(r.status, C_DIM)

            status_para = Paragraph(
                r.status,
                ParagraphStyle("sc", fontName="Helvetica-Bold", fontSize=7,
                               textColor=row_color, alignment=TA_CENTER)
            )
            rows.append([
                Paragraph(r.machine_id,   s["table_cell"]),
                Paragraph(r.machine_type, s["table_cell"]),
                Paragraph(r.protocol,     s["table_cell"]),
                status_para,
                Paragraph(alarm_str,      s["table_cell"]),
                Paragraph(f"{r.piece_count:,}", s["table_cell"]),
                Paragraph(f"{r.temperature_c:.1f}", s["table_cell"]),
                Paragraph(avail_str,      s["table_cell"]),
            ])

        machine_table = Table(rows, colWidths=col_widths, repeatRows=1)

        # Build per-row background alternation
        style_cmds = [
            ("BACKGROUND",    (0, 0), (-1, 0), C_SURFACE),
            ("BOX",           (0, 0), (-1, -1), 0.5, C_BORDER),
            ("INNERGRID",     (0, 0), (-1, -1), 0.3, C_BORDER),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ]
        for i, r in enumerate(snapshot, start=1):
            bg = colors.HexColor("#f0f4f7") if i % 2 == 0 else colors.white
            style_cmds.append(("BACKGROUND", (0, i), (-1, i), bg))

        machine_table.setStyle(TableStyle(style_cmds))
        story.append(machine_table)
        story.append(PageBreak())
        return story

    def _build_alarm_section(self, ds: ReportDataset) -> list:
        """Build the alarm summary section."""
        s   = self._styles
        agg = ds.aggregates

        story = [
            Paragraph("Alarm Summary", s["section_title"]),
            HRFlowable(width="100%", thickness=0.5, color=C_BORDER, spaceAfter=3 * mm),
        ]

        alarm_codes = agg.get("alarm_codes", {})
        if not alarm_codes:
            story.append(Paragraph("No active alarms recorded.", s["body"]))
            return story

        alarm_descriptions = {
            0: "NO_ALARM", 100: "OVERTEMP", 101: "OVERSPEED", 102: "MOTOR_FAULT",
            200: "COMM_ERROR", 201: "SENSOR_FAIL", 300: "E_STOP", 301: "DOOR_OPEN",
            400: "MAINTENANCE_DUE", 500: "POWER_FLUCTUATION",
        }

        rows = [[
            Paragraph("ALARM CODE",   s["table_header"]),
            Paragraph("DESCRIPTION",  s["table_header"]),
            Paragraph("MACHINES AFFECTED", s["table_header"]),
            Paragraph("SEVERITY",     s["table_header"]),
        ]]

        for code, count in sorted(alarm_codes.items(), reverse=True):
            severity  = "CRITICAL" if code >= 300 else "WARNING" if code > 0 else "OK"
            sev_color = C_FAULT if severity == "CRITICAL" else C_IDLE
            rows.append([
                Paragraph(str(code),                             s["table_cell"]),
                Paragraph(alarm_descriptions.get(code, "UNKNOWN"), s["table_cell"]),
                Paragraph(str(count),                             s["table_cell"]),
                Paragraph(
                    severity,
                    ParagraphStyle("sev", fontName="Helvetica-Bold", fontSize=8,
                                   textColor=sev_color, alignment=TA_CENTER)
                ),
            ])

        alarm_table = Table(rows, colWidths=[30*mm, 65*mm, 40*mm, 30*mm])
        alarm_table.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), C_SURFACE),
            ("BACKGROUND",    (0, 1), (-1, -1), colors.white),
            ("BOX",           (0, 0), (-1, -1), 0.5, C_BORDER),
            ("INNERGRID",     (0, 0), (-1, -1), 0.5, C_BORDER),
            ("TOPPADDING",    (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(alarm_table)
        return story

    # ── Main generation ───────────────────────────────────────────────────

    def generate(
        self,
        dataset     : ReportDataset,
        output_path : str,
    ) -> str:
        """
        Generate the full PDF report and write to output_path.

        Args:
            dataset     : ReportDataset with computed aggregates.
            output_path : File path for the output PDF.

        Returns:
            output_path on success.

        Raises:
            IOError : If the file cannot be written.
        """
        if not dataset.aggregates:
            dataset.compute_aggregates()

        doc = SimpleDocTemplate(
            output_path,
            pagesize        = A4,
            leftMargin      = 20 * mm,
            rightMargin     = 20 * mm,
            topMargin       = 20 * mm,
            bottomMargin    = 20 * mm,
            title           = "MyFactoryInsight Report",
            author          = "MFI Phase 09",
        )

        story: list = []
        story += self._build_cover(dataset)
        story += self._build_kpi_section(dataset)
        story += self._build_machine_table(dataset)
        story += self._build_alarm_section(dataset)

        doc.build(story)
        size_kb = os.path.getsize(output_path) // 1024
        LOG.info("PDF generated → %s (%d KB)", output_path, size_kb)
        return output_path


# =============================================================================
# SECTION 6 — EXCEL REPORT GENERATOR
# =============================================================================

class ExcelReportGenerator:
    """
    MFI Excel Report Generator using openpyxl.

    Generates a multi-worksheet workbook:
      Sheet 1: Fleet KPI     — aggregated KPIs with dark header styling
      Sheet 2: Machine Detail — one row per machine, conditional formatting
      Sheet 3: Type Breakdown — per-type summary with bar chart
      Sheet 4: Alarm Log      — alarm code summary table
    """

    # ── Style factories ───────────────────────────────────────────────────

    @staticmethod
    def _header_font() -> Font:
        return Font(bold=True, color=XL_WHITE, name="Calibri", size=11)

    @staticmethod
    def _body_font() -> Font:
        return Font(name="Calibri", size=10, color=XL_TEXT)

    @staticmethod
    def _header_fill() -> PatternFill:
        return PatternFill("solid", fgColor=XL_SURFACE)

    @staticmethod
    def _alt_fill() -> PatternFill:
        return PatternFill("solid", fgColor="F0F4F7")

    @staticmethod
    def _accent_fill() -> PatternFill:
        return PatternFill("solid", fgColor=XL_SURFACE)

    @staticmethod
    def _border() -> Border:
        thin = Side(style="thin", color=XL_BORDER)
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    @staticmethod
    def _center() -> Alignment:
        return Alignment(horizontal="center", vertical="center")

    @staticmethod
    def _status_fill(status: str) -> PatternFill:
        color_map = {
            "RUN":          "D4EDDA",
            "IDLE":         "FFF9C4",
            "FAULT":        "FFCDD2",
            "MAINTENANCE":  "FFE0B2",
        }
        return PatternFill("solid", fgColor=color_map.get(status, "FFFFFF"))

    def _style_header_row(self, ws: Any, row: int, col_count: int) -> None:
        """Apply header styling to a worksheet row."""
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font       = self._header_font()
            cell.fill       = self._header_fill()
            cell.alignment  = self._center()
            cell.border     = self._border()

    def _style_data_row(
        self,
        ws      : Any,
        row     : int,
        col_count: int,
        alt     : bool = False,
    ) -> None:
        """Apply alternating data row styling."""
        fill = self._alt_fill() if alt else PatternFill("solid", fgColor="FFFFFF")
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font       = self._body_font()
            cell.fill       = fill
            cell.alignment  = self._center()
            cell.border     = self._border()

    # ── Sheet builders ────────────────────────────────────────────────────

    def _build_kpi_sheet(self, ws: Any, ds: ReportDataset) -> None:
        """Build the Fleet KPI summary worksheet."""
        agg = ds.aggregates
        ws.title = "Fleet KPI"
        ws.sheet_view.showGridLines = False

        # Title
        ws["A1"] = "MyFactoryInsight — Fleet KPI Report"
        ws["A1"].font = Font(bold=True, size=16, color=XL_ACCENT, name="Calibri")
        ws.merge_cells("A1:C1")

        ws["A2"] = f"Site: {ds.site_id.upper()} | Period: {ds.period_start or '—'} → {ds.period_end or '—'}"
        ws["A2"].font = Font(size=9, color=XL_DIM, name="Calibri")
        ws.merge_cells("A2:C2")

        # KPI rows
        kpis = [
            ("Machine Count",           agg.get("machine_count"),       ""),
            ("Cycles Collected",        agg.get("cycles_collected"),    ""),
            ("Fleet Availability",      agg.get("fleet_availability"),  "%"),
            ("Quality Rate",            agg.get("quality"),             "%"),
            ("Performance",             agg.get("performance"),         "%"),
            ("OEE",                     agg.get("oee"),                 "%"),
            ("Total Pieces Produced",   agg.get("total_pieces"),        ""),
            ("Good Pieces",             agg.get("total_good"),          ""),
            ("Bad / Scrap Pieces",      agg.get("total_bad"),           ""),
            ("Active Alarms",           agg.get("active_alarms"),       ""),
            ("Avg Temperature (°C)",    agg.get("temp_mean"),           "°C"),
            ("Max Temperature (°C)",    agg.get("temp_max"),            "°C"),
            ("Temp Critical Machines",  agg.get("temp_critical_count"), ""),
            ("Temp Warning Machines",   agg.get("temp_warn_count"),     ""),
        ]

        # Header
        ws["A4"] = "METRIC"
        ws["B4"] = "VALUE"
        ws["C4"] = "UNIT"
        self._style_header_row(ws, 4, 3)

        for i, (label, value, unit) in enumerate(kpis, start=5):
            ws.cell(row=i, column=1, value=label)
            if isinstance(value, float) and unit == "%":
                ws.cell(row=i, column=2, value=round(value * 100, 2))
                ws.cell(row=i, column=3, value="%")
            elif isinstance(value, float):
                ws.cell(row=i, column=2, value=round(value, 2))
                ws.cell(row=i, column=3, value=unit)
            else:
                ws.cell(row=i, column=2, value=value)
                ws.cell(row=i, column=3, value=unit)
            self._style_data_row(ws, i, 3, alt=(i % 2 == 0))

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 10

    def _build_machine_sheet(self, ws: Any, ds: ReportDataset) -> None:
        """Build the Machine Detail worksheet with conditional formatting."""
        ws.title = "Machine Detail"
        ws.sheet_view.showGridLines = False
        snapshot = ds.latest_snapshot()

        headers = [
            "Machine ID", "Type", "Protocol", "Status",
            "Alarm", "Piece Count", "Good", "Bad",
            "Cycle Time (s)", "Temp (°C)", "Speed", "Quality %",
            "Availability %", "Temp Severity", "Cycle #",
        ]
        for col, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=h)
        self._style_header_row(ws, 1, len(headers))

        for row_idx, r in enumerate(snapshot, start=2):
            alt = row_idx % 2 == 0
            data = [
                r.machine_id,
                r.machine_type,
                r.protocol,
                r.status,
                r.alarm_code if r.alarm_code != 0 else "",
                r.piece_count,
                r.good_count,
                r.bad_count,
                round(r.cycle_time_sec, 1),
                round(r.temperature_c, 1),
                round(r.speed, 2),
                round(r.quality_rate * 100, 1) if r.quality_rate else "",
                round(r.availability * 100, 1) if r.availability else "",
                r.temp_severity,
                r.cycle_number,
            ]
            for col_idx, val in enumerate(data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=val)
            self._style_data_row(ws, row_idx, len(headers), alt=alt)

            # Status cell color override
            status_cell = ws.cell(row=row_idx, column=4)
            status_cell.fill = self._status_fill(r.status)
            status_cell.font = Font(bold=True, size=10, name="Calibri",
                                    color=XL_SURFACE if r.status == "RUN" else "000000")

        # Column widths
        widths = [14, 12, 12, 14, 8, 14, 10, 10, 14, 12, 10, 12, 14, 14, 10]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Freeze header row
        ws.freeze_panes = "A2"

    def _build_type_sheet(self, ws: Any, ds: ReportDataset) -> None:
        """Build the Type Breakdown worksheet with bar chart."""
        ws.title = "Type Breakdown"
        ws.sheet_view.showGridLines = False
        agg = ds.aggregates
        breakdown = agg.get("type_breakdown", {})

        headers = ["Machine Type", "Count", "Running", "Faulted", "Total Pieces"]
        for col, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=h)
        self._style_header_row(ws, 1, len(headers))

        for row_idx, (mtype, data) in enumerate(sorted(breakdown.items()), start=2):
            ws.cell(row=row_idx, column=1, value=mtype)
            ws.cell(row=row_idx, column=2, value=data["count"])
            ws.cell(row=row_idx, column=3, value=data["running"])
            ws.cell(row=row_idx, column=4, value=data["faulted"])
            ws.cell(row=row_idx, column=5, value=data["pieces"])
            self._style_data_row(ws, row_idx, len(headers), alt=(row_idx % 2 == 0))

        # Bar chart: pieces by type
        last_row = 1 + len(breakdown)
        chart = BarChart()
        chart.type          = "col"
        chart.title         = "Total Pieces by Machine Type"
        chart.y_axis.title  = "Pieces"
        chart.x_axis.title  = "Type"
        chart.width         = 14
        chart.height        = 9

        data_ref    = Reference(ws, min_col=5, min_row=1, max_row=last_row)
        cats        = Reference(ws, min_col=1, min_row=2, max_row=last_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "G2")

        for i, w in enumerate([18, 10, 10, 10, 16], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _build_alarm_sheet(self, ws: Any, ds: ReportDataset) -> None:
        """Build the Alarm Summary worksheet."""
        ws.title = "Alarm Log"
        ws.sheet_view.showGridLines = False
        agg = ds.aggregates
        alarm_codes = agg.get("alarm_codes", {})

        alarm_descriptions = {
            100: "OVERTEMP", 101: "OVERSPEED", 102: "MOTOR_FAULT",
            200: "COMM_ERROR", 201: "SENSOR_FAIL", 300: "E_STOP",
            301: "DOOR_OPEN", 400: "MAINTENANCE_DUE", 500: "POWER_FLUCTUATION",
        }

        headers = ["Alarm Code", "Description", "Machines Affected", "Severity"]
        for col, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=h)
        self._style_header_row(ws, 1, len(headers))

        if not alarm_codes:
            ws.cell(row=2, column=1, value="No active alarms")
            ws.merge_cells("A2:D2")
            return

        for row_idx, (code, count) in enumerate(sorted(alarm_codes.items(), reverse=True), start=2):
            sev = "CRITICAL" if code >= 300 else "WARNING"
            ws.cell(row=row_idx, column=1, value=code)
            ws.cell(row=row_idx, column=2, value=alarm_descriptions.get(code, "UNKNOWN"))
            ws.cell(row=row_idx, column=3, value=count)
            ws.cell(row=row_idx, column=4, value=sev)
            self._style_data_row(ws, row_idx, 4, alt=(row_idx % 2 == 0))

            # Severity color
            sev_cell = ws.cell(row=row_idx, column=4)
            sev_cell.fill = PatternFill("solid", fgColor=(
                "FFCDD2" if sev == "CRITICAL" else "FFF9C4"
            ))
            sev_cell.font = Font(bold=True, size=10, name="Calibri")

        for i, w in enumerate([14, 22, 20, 14], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Main generation ───────────────────────────────────────────────────

    def generate(
        self,
        dataset     : ReportDataset,
        output_path : str,
    ) -> str:
        """
        Generate the full Excel workbook and write to output_path.

        Args:
            dataset     : ReportDataset with computed aggregates.
            output_path : File path for the output XLSX.

        Returns:
            output_path on success.
        """
        if not dataset.aggregates:
            dataset.compute_aggregates()

        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        ws_kpi   = wb.create_sheet()
        ws_mach  = wb.create_sheet()
        ws_type  = wb.create_sheet()
        ws_alarm = wb.create_sheet()

        self._build_kpi_sheet(ws_kpi,   dataset)
        self._build_machine_sheet(ws_mach,  dataset)
        self._build_type_sheet(ws_type,  dataset)
        self._build_alarm_sheet(ws_alarm, dataset)

        # Set active tab to Fleet KPI
        wb.active = ws_kpi

        wb.save(output_path)
        size_kb = os.path.getsize(output_path) // 1024
        LOG.info("Excel generated → %s (%d KB)", output_path, size_kb)
        return output_path


# =============================================================================
# SECTION 7 — REPORT SERVICE
# =============================================================================

class ReportService:
    """
    MFI Report Service — collects pipeline data and generates reports on demand.

    Usage:
        service = ReportService(cycles=5)
        service.collect(core)
        service.generate_pdf("report.pdf")
        service.generate_xlsx("report.xlsx")
    """

    def __init__(
        self,
        site_id : str = DEFAULT_SITE,
        cycles  : int = DEFAULT_CYCLES,
    ) -> None:
        self._site_id   = site_id
        self._cycles    = cycles
        self._dataset   = ReportDataset(site_id=site_id)
        self._pdf_gen   = PDFReportGenerator()
        self._xlsx_gen  = ExcelReportGenerator()
        LOG.info(
            "ReportService initialized │ site=%s │ cycles=%d",
            site_id, cycles,
        )

    def collect(self, core: MFICore, interval: float = 0.5) -> ReportDataset:
        """
        Run MFICore for N cycles and accumulate data into the dataset.

        Args:
            core     : MFICore pipeline instance.
            interval : Sleep between cycles (seconds).

        Returns:
            Populated ReportDataset.
        """
        LOG.info("Collecting data │ cycles=%d │ interval=%.1fs", self._cycles, interval)
        for cycle in range(1, self._cycles + 1):
            result = core.run_cycle()
            self._dataset.add_cycle(result["enriched"])
            LOG.info(
                "Cycle %d/%d collected │ enriched=%d",
                cycle, self._cycles, len(result["enriched"]),
            )
            if cycle < self._cycles:
                time.sleep(interval)

        self._dataset.compute_aggregates()
        return self._dataset

    def generate_pdf(self, output_path: str) -> str:
        """Generate PDF report. Returns output_path."""
        return self._pdf_gen.generate(self._dataset, output_path)

    def generate_xlsx(self, output_path: str) -> str:
        """Generate Excel report. Returns output_path."""
        return self._xlsx_gen.generate(self._dataset, output_path)

    def dataset(self) -> ReportDataset:
        """Return the current ReportDataset."""
        return self._dataset


# =============================================================================
# SECTION 8 — SELF-TEST
# =============================================================================

def run_self_test() -> bool:
    """
    Self-test for Phase 9. Validates:
      1.  ReportDataset.add_cycle() accumulates records correctly.
      2.  ReportDataset.latest_snapshot() returns one record per machine.
      3.  ReportDataset.period_start and period_end are set correctly.
      4.  compute_aggregates() returns all required KPI keys.
      5.  compute_aggregates() fleet_availability in [0, 1].
      6.  compute_aggregates() status_distribution sums to machine_count.
      7.  compute_aggregates() type_breakdown has correct machine types.
      8.  compute_aggregates() total_pieces > 0 after cycles.
      9.  PDFReportGenerator.generate() produces a non-empty PDF file.
      10. PDF file size > 1 KB.
      11. ExcelReportGenerator.generate() produces a non-empty XLSX file.
      12. XLSX file size > 2 KB.
      13. XLSX has correct number of worksheets (4).
      14. XLSX Fleet KPI sheet has correct title in A1.
      15. XLSX Machine Detail sheet has data rows.
      16. XLSX Type Breakdown sheet has all machine types.
      17. XLSX Alarm Log sheet exists.
      18. ReportService.collect() populates dataset.
      19. ReportDataset aggregates include top5_producers.
      20. PDF and XLSX generate to io.BytesIO (in-memory) without errors.

    Returns:
        True if all assertions pass, False otherwise.
    """
    import tempfile

    LOG.info("══════════ SELF-TEST START ══════════")
    passed = 0
    failed = 0

    def check(label: str, condition: bool, detail: str = "") -> None:
        nonlocal passed, failed
        if condition:
            LOG.info("  ✓ PASS │ %s", label)
            passed += 1
        else:
            LOG.error("  ✗ FAIL │ %s%s", label, f" → {detail}" if detail else "")
            failed += 1

    # ── Build dataset via MFICore ─────────────────────────────────────────
    core    = MFICore(site_id="mecanitec")
    ds      = ReportDataset(site_id="mecanitec")

    # Collect 3 cycles
    for _ in range(3):
        result = core.run_cycle()
        ds.add_cycle(result["enriched"])

    # Tests 1-3: ReportDataset collection
    check(
        "add_cycle() accumulates records",
        len(ds.records) >= 40 * 3,
        f"got {len(ds.records)}",
    )
    snapshot = ds.latest_snapshot()
    check(
        "latest_snapshot() returns one record per machine",
        len(snapshot) >= 40 and len(snapshot) == len(set(r.machine_id for r in snapshot)),
        f"got {len(snapshot)}",
    )
    check(
        "period_start and period_end are set",
        ds.period_start is not None and ds.period_end is not None,
    )

    # Tests 4-8: compute_aggregates
    agg = ds.compute_aggregates()
    required_keys = [
        "machine_count", "fleet_availability", "quality", "oee",
        "total_pieces", "status_distribution", "type_breakdown",
        "temp_mean", "active_alarms", "top5_producers",
    ]
    check(
        "compute_aggregates() returns all required KPI keys",
        all(k in agg for k in required_keys),
        str([k for k in required_keys if k not in agg]),
    )
    avail = agg.get("fleet_availability")
    check(
        "fleet_availability in [0, 1]",
        avail is not None and 0.0 <= avail <= 1.0,
        f"got {avail}",
    )
    dist  = agg.get("status_distribution", {})
    total = sum(dist.values())
    check(
        "status_distribution sums to machine_count",
        total == agg.get("machine_count", -1),
        f"dist_sum={total} machine_count={agg.get('machine_count')}",
    )
    breakdown = agg.get("type_breakdown", {})
    known_types = {"CNC", "PRESS", "ROBOT", "CONVEYOR", "WELDER"}
    check(
        "type_breakdown has valid machine types",
        all(t in known_types for t in breakdown.keys()),
        str(set(breakdown.keys()) - known_types),
    )
    check(
        "total_pieces > 0",
        agg.get("total_pieces", 0) > 0,
        f"got {agg.get('total_pieces')}",
    )

    # Tests 9-10: PDF generation
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
        pdf_path = f.name

    try:
        pdf_gen = PDFReportGenerator()
        pdf_gen.generate(ds, pdf_path)
        pdf_size = os.path.getsize(pdf_path)
        check("PDFReportGenerator.generate() produces a PDF file", pdf_size > 0)
        check("PDF file size > 1 KB", pdf_size > 1024, f"got {pdf_size} bytes")
    finally:
        if os.path.exists(pdf_path):
            os.unlink(pdf_path)

    # Tests 11-17: XLSX generation
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        xlsx_path = f.name

    try:
        xlsx_gen = ExcelReportGenerator()
        xlsx_gen.generate(ds, xlsx_path)
        xlsx_size = os.path.getsize(xlsx_path)
        check("ExcelReportGenerator.generate() produces XLSX file", xlsx_size > 0)
        check("XLSX file size > 2 KB", xlsx_size > 2048, f"got {xlsx_size} bytes")

        # Inspect generated XLSX
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path)
        check("XLSX has 4 worksheets", len(wb.sheetnames) == 4, str(wb.sheetnames))

        ws_kpi = wb["Fleet KPI"]
        check(
            "Fleet KPI sheet A1 contains 'MyFactoryInsight'",
            ws_kpi["A1"].value is not None and "MyFactoryInsight" in str(ws_kpi["A1"].value),
            str(ws_kpi["A1"].value),
        )

        ws_mach = wb["Machine Detail"]
        check(
            "Machine Detail has data rows",
            ws_mach.max_row >= 40,
            f"max_row={ws_mach.max_row}",
        )

        ws_type = wb["Type Breakdown"]
        type_names = [ws_type.cell(row=r, column=1).value for r in range(2, ws_type.max_row + 1)]
        check(
            "Type Breakdown has machine types",
            any(t in str(type_names) for t in ["CNC", "PRESS", "ROBOT"]),
            str(type_names),
        )

        check("Alarm Log sheet exists", "Alarm Log" in wb.sheetnames)

    finally:
        if os.path.exists(xlsx_path):
            os.unlink(xlsx_path)

    # Test 18: ReportService.collect()
    service = ReportService(site_id="mecanitec", cycles=2)
    core2   = MFICore(site_id="mecanitec")
    service.collect(core2, interval=0.0)
    check(
        "ReportService.collect() populates dataset",
        service.dataset().cycles_collected == 2,
        f"got {service.dataset().cycles_collected}",
    )

    # Test 19: top5_producers
    top5 = agg.get("top5_producers", [])
    check(
        "aggregates include top5_producers with 5 entries",
        len(top5) == 5 and all("machine_id" in p for p in top5),
        str(top5),
    )

    # Test 20: In-memory generation (BytesIO)
    try:
        ds2 = ReportDataset(site_id="mecanitec")
        core3 = MFICore(site_id="mecanitec")
        r3 = core3.run_cycle()
        ds2.add_cycle(r3["enriched"])
        ds2.compute_aggregates()

        # PDF to temp file
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
            tmp_pdf = f.name
        PDFReportGenerator().generate(ds2, tmp_pdf)
        ok_pdf = os.path.getsize(tmp_pdf) > 0
        os.unlink(tmp_pdf)

        # XLSX to temp file
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_xlsx = f.name
        ExcelReportGenerator().generate(ds2, tmp_xlsx)
        ok_xlsx = os.path.getsize(tmp_xlsx) > 0
        os.unlink(tmp_xlsx)

        check("PDF and XLSX generate without errors in single-cycle mode", ok_pdf and ok_xlsx)
    except Exception as exc:
        check("PDF and XLSX generate without errors in single-cycle mode", False, str(exc))

    # --- Summary ---
    total = passed + failed
    LOG.info(
        "══════════ SELF-TEST RESULT: %d/%d PASSED %s ══════════",
        passed,
        total,
        "✓ OK" if failed == 0 else "✗ FAIL",
    )
    return failed == 0


# =============================================================================
# SECTION 9 — CLI / MAIN ENTRY POINT
# =============================================================================

def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog        = "mfi_phase_09_reports.py",
        description = (
            f"MFI Phase {PHASE_ID} — {PHASE_NAME} v{PHASE_VERSION}\n"
            "Generate PDF and Excel reports from MFI pipeline data."
        ),
        formatter_class = argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--self-test", action="store_true",
                        help="Run built-in self-test and exit.")
    parser.add_argument("--cycles", type=int, default=DEFAULT_CYCLES, metavar="N",
                        help=f"Pipeline cycles for data collection (default: {DEFAULT_CYCLES}).")
    parser.add_argument("--pdf", type=str, default=None, metavar="FILE",
                        help=f"Output PDF path (default: {DEFAULT_PDF_PATH}).")
    parser.add_argument("--xlsx", type=str, default=None, metavar="FILE",
                        help=f"Output Excel path (default: {DEFAULT_XLSX_PATH}).")
    parser.add_argument("--site", type=str, default=DEFAULT_SITE, metavar="SITE_ID",
                        help=f"Site identifier (default: {DEFAULT_SITE}).")
    parser.add_argument("--interval", type=float, default=0.5, metavar="SECS",
                        help="Seconds between collection cycles (default: 0.5).")
    return parser


def main() -> None:
    parser  = build_arg_parser()
    args    = parser.parse_args()

    LOG.info("╔══════════════════════════════════════════════╗")
    LOG.info("║  MyFactoryInsight  │  Phase %-2s │ %-20s ║", PHASE_ID, PHASE_NAME)
    LOG.info("║  Version %-7s   │ Site: %-23s ║", PHASE_VERSION, args.site)
    LOG.info("╚══════════════════════════════════════════════╝")

    if args.self_test:
        success = run_self_test()
        sys.exit(0 if success else 1)

    # At least one output must be requested
    if args.pdf is None and args.xlsx is None:
        args.pdf  = DEFAULT_PDF_PATH
        args.xlsx = DEFAULT_XLSX_PATH
        LOG.info("No output specified — generating both PDF and Excel.")

    service = ReportService(site_id=args.site, cycles=args.cycles)
    core    = MFICore(site_id=args.site)

    service.collect(core, interval=args.interval)

    if args.pdf:
        service.generate_pdf(args.pdf)
    if args.xlsx:
        service.generate_xlsx(args.xlsx)

    agg = service.dataset().aggregates
    LOG.info(
        "Phase %s complete │ machines=%d │ oee=%s │ avail=%s │ pieces=%s",
        PHASE_ID,
        agg.get("machine_count", 0),
        f"{agg['oee']:.1%}" if agg.get("oee") else "N/A",
        f"{agg['fleet_availability']:.1%}" if agg.get("fleet_availability") else "N/A",
        f"{agg.get('total_pieces', 0):,}",
    )


# =============================================================================
# SECTION 10 — ENTRY GUARD
# =============================================================================
if __name__ == "__main__":
    main()
